from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
import os
from werkzeug.security import generate_password_hash, check_password_hash
from flask_cors import CORS
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)
CORS(app)

excel_file = "users.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(excel_file):
    logger.info(f"Creating new Excel file: {excel_file}")
    wb = Workbook()
    ws = wb.active
    ws.append(["Username", "Password"])  # headers
    wb.save(excel_file)

@app.route('/')
def home():
    logger.debug("Home route accessed")
    return "Login API Server is running! Available endpoints: /login, /signup, /test"

@app.route('/test')
def test():
    logger.debug("Test route accessed")
    return jsonify({"status": "success", "message": "Server is running!"})

@app.route('/signup', methods=['POST'])
def signup():
    logger.debug(f"Signup request received: {request.json}")
    try:
        data = request.json
        username = data.get("username")
        password = data.get("password")

        if not username or not password:
            logger.warning("Signup failed: Missing username or password")
            return jsonify({"success": False, "message": "Username and password are required"}), 400

        wb = load_workbook(excel_file)
        ws = wb.active

        # Check for existing username
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                logger.warning(f"Signup failed: Username {username} already exists")
                return jsonify({"success": False, "message": "Username already exists"}), 400

        # Hash the password before saving
        hashed_password = generate_password_hash(password)

        # Save user
        ws.append([username, hashed_password])
        wb.save(excel_file)
        logger.info(f"User {username} signed up successfully")
        return jsonify({"success": True, "message": "Signup successful"}), 201
    except Exception as e:
        logger.error(f"Error in signup: {str(e)}")
        return jsonify({"success": False, "message": "Server error occurred"}), 500

@app.route('/login', methods=['POST'])
def login():
    logger.debug(f"Login request received: {request.json}")
    try:
        data = request.json
        username = data.get("username")
        password = data.get("password")

        if not username or not password:
            logger.warning("Login failed: Missing username or password")
            return jsonify({"success": False, "message": "Username and password are required"}), 400

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and check_password_hash(row[1], password):
                logger.info(f"User {username} logged in successfully")
                return jsonify({"success": True, "message": "Login successful"}), 200

        logger.warning(f"Login failed: Invalid credentials for user {username}")
        return jsonify({"success": False, "message": "Invalid username or password"}), 401
    except Exception as e:
        logger.error(f"Error in login: {str(e)}")
        return jsonify({"success": False, "message": "Server error occurred"}), 500

if __name__ == '__main__':
    logger.info("Starting Login API Server...")
    logger.info("Available endpoints:")
    logger.info("- http://localhost:5000/ (Home)")
    logger.info("- http://localhost:5000/test (Test endpoint)")
    logger.info("- http://localhost:5000/login (POST - Login)")
    logger.info("- http://localhost:5000/signup (POST - Signup)")
    app.run(debug=True, host='0.0.0.0', port=5000) 